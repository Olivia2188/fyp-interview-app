from flask import Flask, render_template, request, jsonify, send_from_directory
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
import os
import json     #to read, write & manipulate JSON data
import random
import requests
app = Flask(__name__)

def evaluate_with_AI(answer, rubric, maxMark):
    endpoint = "https://openrouter.ai/api/v1/chat/completions" #endpoint: whr u send  request to(API URL)
    api_token = "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"

    prompt = (
        f"Rubric: {rubric}\n"
        f"Max mark: {maxMark}\n"
        f"Student's answer: {answer}\n"
        f"Based on the rubric and max mark, give a score and a reason."
        f"Keep the score within the max mark. Format: 'Score: X/Y | Reason: ... '"
    )
    headers = { #include extra information for server
        "Content-Type": "application/json",
        "Authorization": f"Bearer {api_token}" #tell API who am I, use the token to log in, not allowed to use if dk who am I 
    }
    payload = { #actual data sending to AI
        "model": "mistralai/mistral-7b-instruct",
        "messages":[  #tells AI who's speaking n what they speaking(content)
            {"role": "system","content": "You are an assistant that grades answers based on the rubrics."}, #setting the environment, telling the role
            {"role": "user", "content": prompt} #user is the one giving instructions, send in instructions under 'content'
        ]
    }

    response = requests.post(endpoint, headers=headers, json=payload) #sends HTTP POST request

    if response.status_code == 200: #success(ok)
        return response.json()["choices"][0]["message"]["content"]
    else:
        print("AI Evaluation Error:", response.status_code, response.text)
        return "Evaluation failed."
        
@app.route('/download/<filename>')
def download(filename):
    return send_from_directory('generated_reports', filename, as_attachment=True)

@app.route('/uploadQuestion', methods=['POST'])
def uploadQuestion():
    file = request.files.get('file') # capture files sent from frontend that named 'file'

    if not file:
        return jsonify({"error": "No file uploaded"}), 400 # if file doesn't exist, return an error msg in JSON
    
    try:
        wb = load_workbook(filename = file) # workbook = open & read workbook with the filename called file
        sheet = wb.active # sheet = the first sheet/ last saved in the wb 

        expected_headers = ["Question Text", "Rubric", "Time Given(min)", "Awarded Marks"]
        actual_headers = [cell.value for cell in sheet[1]] # take the text inside the cell for row 1

        if expected_headers != actual_headers: # if they are different
            return jsonify({
                "error": f"Could not import from this spreadsheet. Please follow the sample template."
            }), 400
        
        questions = [] #empty list for storage

        for idx, row in enumerate(sheet.iter_rows(min_row = 2, values_only= True), start = 2): # going thru the wb row by row start from row 2, take the data in the cells
            if not any(row): # all 0 or empty (false), skip the row
                continue 

            question_text, rubric, time_given, awarded_marks = row # unpacking the row into 3 variables
            if question_text is None or rubric is None or time_given is None or awarded_marks is None:
                return jsonify({
                    "error": f"Missing the data in row {idx}. All fields are required" # idx count how many loops have happened like which excel row I'm at
                }), 400
            
            try:
                time_given = float(time_given)
                if time_given < 0:
                    raise ValueError
            except:
                return jsonify({"error": f"Invalid in row {idx}. Invalid time input."}), 400
            
            try:
                awarded_marks = float(awarded_marks)
                if awarded_marks < 0:
                    raise ValueError
            except:
                return jsonify({"error": f"Invalid in row {idx}. Invalid marks input ."}), 400
            
            questions.append({ # add the data to the empty list
                "text": str(question_text).strip(),
                "rubric": str(rubric).strip(),
                "time_given": time_given,
                "marks": awarded_marks
            })

            with open("uploadQuestion.json", "w") as f: #put the qns list inside the file called uploadQuestion for later use
                json.dump(questions, f)

        if not questions:
            return jsonify({"error": "No valid questions found in the file."}), 400
        
        return jsonify({"questions": questions}) # convert back to JSON format and sned back to the JS
    
    except Exception as e:
        return jsonify({"error":f"Failed to process Excel file:{str(e)}"}), 500

@app.route('/start_interview', methods=['GET'])
def start_interview():
    try:
        with open("uploadQuestion.json", "r") as f: #open the file in read mode
            questions = json.load(f) #load the JSON data from f, convert to python data and save it under questions
        random.shuffle(questions) 
        return jsonify({"questions":questions}) #convert python data to JSON data and save under the key named questions
    except Exception as e:
        return jsonify({"error": f"Failed to load questions: {str(e)}"}), 500

@app.route('/save_transcription', methods=['POST']) #set up a route to go the blk of codes, telling server fetch uses POST to send data
def save_transcription():
    data = request.get_json() #extract json data sent from fetch & turns it in python dictionary
    text = data.get("text", "") # extract the data in text, leave it empty if it's empty
    studentName = data.get("studentName")
    admin = data.get("admin")
    module = data.get("module")
    question = data.get("question")
    index = data.get("index")

    if not text.strip(): #strip remove leading/trailing spaces--> if text empty after removing spaces, carry actions
        return jsonify({"message": "No text to save!"}), 400  #400 means HTTP error (bad request, invalid input, users fault)
    
    filename = f"{admin}_{module}_{studentName}.xlsx" #name of the excel file"
    folder = "generated_reports" #folder to store all the Excel files
    os.makedirs(folder, exist_ok=True) # creates the folder if it doesn't exist
    excel_path = os.path.join(folder, filename) # creates the correct full path string

    if os.path.exists(excel_path):
        wb = load_workbook(excel_path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
    
        ws["A1"] = "Student Name"
        ws["B1"] = studentName
        ws["A2"] = "Admin No"
        ws["B2"] = admin
        ws["A3"] = "Module Code"
        ws["B3"] = module

        headers = ["Serial", "Question", "Response", "Score", "Remarks"]
        ws.append([])
        ws.append(headers)

        for col_num in range(1, len(headers)+1):
            cell = ws.cell(row=5, column=col_num)
            cell.font = Font(bold=True)

    target_row = int(index) + 6 #calculating row: index 0 = row 6, index 1 = row 7...
    serial = int(index) + 1

    
    with open("uploadQuestion.json", "r") as f: #open the file in read mode(comes from /uploadQuestion)
        allQuestions = json.load(f) #load the JSON data from f, convert to python dic and save it under allQuestions

    currentQn = allQuestions[int(index)]
    rubric = currentQn["rubric"]
    maxMark = currentQn["marks"]

    aiResult = evaluate_with_AI(text, rubric, maxMark) #text is student transcribed text(from this func), rubric & maxMark alr declare on top

    try:
        score_part, reason_part = aiResult.split("|")
        score = score_part.strip().replace("Score:","").split("/")[0].strip()
        reason = reason_part.strip().replace("Reason:", "").strip()
    except:
        score = 0
        reason = "Unable to parse AI result"

    ws.cell(row=target_row, column=1, value=serial)   #serial no.
    ws.cell(row=target_row, column=2, value=question)   #question
    ws.cell(row=target_row, column=3, value=text)    #transcribed text
    ws.cell(row=target_row, column=4, value=score)   #Score
    ws.cell(row=target_row, column=5, value=reason)  #rationale behind


    wb.save(excel_path) # save excel to the correct folder

    return jsonify({"message": f"Save successfully as {excel_path}"}) #send response back to frontend

@app.route('/save_questionInfo', methods=["POST"])
def save_questionInfo():
    data = request.get_json()
    title = data.get("title")
    code = data.get("code")

    print(f"Question Info received: {code}:{title}")

    return jsonify({"success": True, "message": "Student info saved!"})

@app.route('/student')
def student():
    return render_template('student.html')

@app.route('/studentInfo')
def studentInfo():
    return render_template('studentInfo.html')

@app.route('/questionInfo')
def questionInfo():
    return render_template('questionInfo.html')

@app.route('/inputQuestion')
def inputQuestion():
    return render_template('inputQuestion.html')

@app.route('/qnMethod')
def qnMethod():
    return render_template('teacher-createQn-Method.html')

@app.route('/createfromScratch')
def createfromScratch():
    return render_template('teacher-createQn-scratch.html')

@app.route('/')
def index():
    return(render_template('index.html'))

@app.route('/save_student', methods=['POST'])
def save_student():
    data = request.get_json()
    studentName = data.get("studentName")
    admin = data.get("admin")
    module = data.get("module")

    print(f"Student Info received: {studentName}, {admin}, {module}")

    return jsonify({"success": True, "message": "Student info saved!"})

if __name__ == '__main__':
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=True)

